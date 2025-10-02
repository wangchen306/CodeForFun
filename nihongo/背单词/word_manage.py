# from openpyxl import load_workbook, Workbook
import openpyxl
import pandas as pd

from utils import split_kanji_kana


class WordManager:
    def __init__(self, wordFilePath, infoFilePath):
        self.fileMaxRow, self.fileMaxCol = 0, 0
        self.wordFilePath, self.infoFilePath = wordFilePath, infoFilePath
        self.init_wordDF(infoFilePath)
        # self.word_df = pd.DataFrame(columns=['kanji', 'kana', 'wrong', 'choose', 'ratio'])
        self.read_excel(wordFilePath)

    def init_wordDF(self, file_path):
        try:
            self.word_df = pd.read_csv(file_path, index_col=0)
        except FileNotFoundError:
            self.word_df = pd.DataFrame(columns=['kanji', 'kana', 'wrong', 'choose', 'ratio'])

    def read_excel(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb['音読み']  # 获取活动工作表
            self.fileMaxRow, self.fileMaxCol = sheet.max_row, sheet.max_column
            for row in range(1, self.fileMaxRow + 1):
                for col in range(3, self.fileMaxCol + 1):
                    word = sheet.cell(row=row, column=col).value
                    kanji_list, kana_list = split_kanji_kana(word)
                    if kanji_list and kana_list:
                        if kanji_list[0] in self.word_df.index:
                            continue
                        else:
                            self.word_df.loc[kanji_list[0]] = [kanji_list[0], kana_list[0], 0, 0, 1]
                    else:
                        continue
        except FileNotFoundError:
            print(f"未找到文件！")

    def save_words(self):
        self.word_df.to_csv(self.infoFilePath)


# ===== 使用示例 =====
if __name__ == "__main__":
    wf = WordManager(wordFilePath="単語.xlsx", infoFilePath='wordsInfo-backup.csv')
    print()
    wf.save_words()